"""Import/export livestock data as CSV or Excel (openpyxl)."""

import csv
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.models.livestock import AttributeDefinition, Livestock, LivestockSex, LivestockStatus, LivestockType
from app.services.attribute import get_attribute_definitions, get_cattle_type
from app.services.livestock import _next_reg_number


_BASE_COLUMNS = [
    "registration_number",
    "name",
    "sex",
    "status",
    "farm_id",
    "is_available_for_breeding",
]

_BOOL_TRUE = {"true", "yes", "1", "t", "y"}


def _to_bool(val: str) -> bool:
    return val.strip().lower() in _BOOL_TRUE


def _normalize_row(row: dict, attr_defs: list[AttributeDefinition]) -> dict:
    """Parse a raw string row into typed fields + attributes dict."""
    attrs = {}
    for defn in attr_defs:
        key = f"attr_{defn.attribute_key}"
        if key in row and row[key] not in ("", None):
            attrs[defn.attribute_key] = row[key].strip()

    return {
        "registration_number": (row.get("registration_number") or "").strip() or None,
        "name": (row.get("name") or "").strip() or None,
        "sex": (row.get("sex") or "FEMALE").strip().upper(),
        "status": (row.get("status") or "ACTIVE").strip().upper(),
        "farm_id": int(row["farm_id"]) if row.get("farm_id") else None,
        "is_available_for_breeding": _to_bool(row.get("is_available_for_breeding", "true")),
        "attributes": attrs,
    }


async def import_livestock_csv(
    content: bytes,
    db: AsyncSession,
    jzd_id: int,
    preview: bool = False,
) -> dict:
    """
    Import livestock from CSV bytes.
    Returns {"created": N, "errors": [...], "preview": [...]} .
    When preview=True no DB writes are done.
    """
    cattle_type = await get_cattle_type(db)
    attr_defs = await get_attribute_definitions(cattle_type.id, db) if cattle_type else []

    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    created = 0
    errors: list[dict] = []
    preview_rows: list[dict] = []

    for i, raw in enumerate(reader, start=2):  # row 1 = header
        try:
            parsed = _normalize_row(raw, attr_defs)
        except Exception as exc:
            errors.append({"row": i, "error": str(exc)})
            continue

        if preview:
            preview_rows.append({"row": i, **parsed})
            continue

        if not parsed["farm_id"]:
            errors.append({"row": i, "error": "farm_id is required"})
            continue

        reg = parsed["registration_number"]
        if not reg:
            reg = await _next_reg_number(jzd_id, db)

        animal = Livestock(
            jzd_id=jzd_id,
            farm_id=parsed["farm_id"],
            registration_number=reg,
            name=parsed["name"],
            sex=LivestockSex(parsed["sex"]),
            status=LivestockStatus(parsed["status"]),
            is_available_for_breeding=parsed["is_available_for_breeding"],
            livestock_type_id=cattle_type.id if cattle_type else None,
            attributes=parsed["attributes"],
        )
        db.add(animal)
        created += 1

    if not preview and created:
        await db.commit()

    return {"created": created, "errors": errors, "preview": preview_rows}


async def import_livestock_excel(
    content: bytes,
    db: AsyncSession,
    jzd_id: int,
    preview: bool = False,
) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "errors": [], "preview": []}

    headers = [str(h).strip() if h else "" for h in rows[0]]

    cattle_type = await get_cattle_type(db)
    attr_defs = await get_attribute_definitions(cattle_type.id, db) if cattle_type else []

    created = 0
    errors: list[dict] = []
    preview_rows: list[dict] = []

    for i, row_vals in enumerate(rows[1:], start=2):
        raw = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row_vals)}
        try:
            parsed = _normalize_row(raw, attr_defs)
        except Exception as exc:
            errors.append({"row": i, "error": str(exc)})
            continue

        if preview:
            preview_rows.append({"row": i, **parsed})
            continue

        if not parsed["farm_id"]:
            errors.append({"row": i, "error": "farm_id is required"})
            continue

        reg = parsed["registration_number"]
        if not reg:
            reg = await _next_reg_number(jzd_id, db)

        animal = Livestock(
            jzd_id=jzd_id,
            farm_id=parsed["farm_id"],
            registration_number=reg,
            name=parsed["name"],
            sex=LivestockSex(parsed["sex"]),
            status=LivestockStatus(parsed["status"]),
            is_available_for_breeding=parsed["is_available_for_breeding"],
            livestock_type_id=cattle_type.id if cattle_type else None,
            attributes=parsed["attributes"],
        )
        db.add(animal)
        created += 1

    if not preview and created:
        await db.commit()

    return {"created": created, "errors": errors, "preview": preview_rows}


async def export_livestock_csv(db: AsyncSession, jzd_id: int) -> bytes:
    cattle_type = await get_cattle_type(db)
    attr_defs = await get_attribute_definitions(cattle_type.id, db) if cattle_type else []

    q = select(Livestock).where(Livestock.jzd_id == jzd_id).options(selectinload(Livestock.farm))
    animals = list((await db.execute(q)).scalars().all())

    out = io.StringIO()
    attr_keys = [d.attribute_key for d in attr_defs]
    fieldnames = _BASE_COLUMNS + [f"attr_{k}" for k in attr_keys]
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()

    for a in animals:
        row = {
            "registration_number": a.registration_number,
            "name": a.name or "",
            "sex": a.sex.value,
            "status": a.status.value,
            "farm_id": a.farm_id or "",
            "is_available_for_breeding": str(a.is_available_for_breeding).lower(),
        }
        attrs = a.attributes or {}
        for k in attr_keys:
            row[f"attr_{k}"] = attrs.get(k, "")
        writer.writerow(row)

    return out.getvalue().encode("utf-8-sig")


async def export_livestock_excel(db: AsyncSession, jzd_id: int) -> bytes:
    cattle_type = await get_cattle_type(db)
    attr_defs = await get_attribute_definitions(cattle_type.id, db) if cattle_type else []

    q = select(Livestock).where(Livestock.jzd_id == jzd_id).options(selectinload(Livestock.farm))
    animals = list((await db.execute(q)).scalars().all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livestock"

    attr_keys = [d.attribute_key for d in attr_defs]
    attr_names = [d.attribute_name for d in attr_defs]
    headers = _BASE_COLUMNS + [f"attr_{k}" for k in attr_keys]
    display_headers = ["Registration Number", "Name", "Sex", "Status", "Farm ID", "Available for Breeding"] + attr_names

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, label in enumerate(display_headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, a in enumerate(animals, start=2):
        attrs = a.attributes or {}
        values = [
            a.registration_number,
            a.name or "",
            a.sex.value,
            a.status.value,
            a.farm_id or "",
            str(a.is_available_for_breeding).lower(),
        ] + [attrs.get(k, "") for k in attr_keys]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_num, column=col, value=val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def get_import_template_excel(db: AsyncSession) -> bytes:
    cattle_type = await get_cattle_type(db)
    attr_defs = await get_attribute_definitions(cattle_type.id, db) if cattle_type else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    attr_keys = [d.attribute_key for d in attr_defs]
    attr_names = [d.attribute_name for d in attr_defs]
    display_headers = ["Registration Number", "Name", "Sex", "Status", "Farm ID", "Available for Breeding"] + attr_names
    col_keys = _BASE_COLUMNS + [f"attr_{k}" for k in attr_keys]

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, (label, key) in enumerate(zip(display_headers, col_keys), start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font

    # Notes row
    notes_row = [
        "CZ + 12 digits or leave blank for auto-generate",
        "Optional name",
        "FEMALE or MALE",
        "ACTIVE / INACTIVE / SOLD / DECEASED",
        "Numeric farm ID (required)",
        "true or false",
    ] + [d.unit or "" for d in attr_defs]
    note_font = Font(italic=True, color="777777")
    for col, note in enumerate(notes_row, start=1):
        cell = ws.cell(row=2, column=col, value=note)
        cell.font = note_font

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
